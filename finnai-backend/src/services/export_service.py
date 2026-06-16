from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font

from models.transaction import Transaction

ExportRow = Transaction | dict


@dataclass(frozen=True)
class ExportFile:
    filename: str
    content_type: str
    content: bytes


def _row_value(row: ExportRow, key: str, attr: str | None = None) -> object:
    if isinstance(row, dict):
        return row.get(key, "")
    return getattr(row, attr or key, "")


class ExportService:
    def export_csv(self, *, rows: list[ExportRow], filename_prefix: str) -> ExportFile:
        output = io.StringIO()
        writer = csv.writer(output)
        has_projected = any(isinstance(r, dict) and "is_projected" in r for r in rows)
        headers = [
            "transaction_date",
            "type",
            "amount_cents",
            "description",
            "notes",
            "account_name",
            "category_name",
        ]
        if has_projected:
            headers.append("is_projected")
        writer.writerow(headers)
        for tx in rows:
            date_val = _row_value(tx, "transaction_date")
            date_str = date_val.isoformat() if hasattr(date_val, "isoformat") else str(date_val)
            row_data = [
                date_str,
                _row_value(tx, "type"),
                int(_row_value(tx, "amount_cents")),
                _row_value(tx, "description"),
                _row_value(tx, "notes") or "",
                _row_value(tx, "account_name"),
                _row_value(tx, "category_name"),
            ]
            if has_projected:
                row_data.append(_row_value(tx, "is_projected") if isinstance(tx, dict) else False)
            writer.writerow(row_data)

        # UTF-8 with BOM for Excel
        content = output.getvalue().encode("utf-8-sig")
        return ExportFile(
            filename=f"{filename_prefix}.csv",
            content_type="text/csv; charset=utf-8",
            content=content,
        )

    def export_xlsx(self, *, rows: list[ExportRow], filename_prefix: str) -> ExportFile:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        has_projected = any(isinstance(r, dict) and "is_projected" in r for r in rows)
        headers = [
            "Transaction Date",
            "Type",
            "Amount (cents)",
            "Description",
            "Notes",
            "Account",
            "Category",
        ]
        if has_projected:
            headers.append("Is Projected")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        total_income = 0
        total_expense = 0

        for tx in rows:
            amount = int(_row_value(tx, "amount_cents"))
            tx_type = str(_row_value(tx, "type"))
            if tx_type == "income":
                total_income += amount
            elif tx_type == "expense":
                total_expense += amount

            date_val = _row_value(tx, "transaction_date")
            date_str = date_val.isoformat() if hasattr(date_val, "isoformat") else str(date_val)
            row_data = [
                date_str,
                tx_type,
                amount,
                _row_value(tx, "description"),
                _row_value(tx, "notes") or "",
                _row_value(tx, "account_name"),
                _row_value(tx, "category_name"),
            ]
            if has_projected:
                row_data.append(_row_value(tx, "is_projected") if isinstance(tx, dict) else False)
            ws.append(row_data)

        ws.append([])
        ws.append(["Totals", "", "", "", "", "", ""])
        ws.append(["Total income (cents)", total_income, "", "", "", "", ""])
        ws.append(["Total expense (cents)", total_expense, "", "", "", "", ""])
        ws.append(["Net (cents)", total_income - total_expense, "", "", "", "", ""])
        ws["A" + str(ws.max_row - 3)].font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        return ExportFile(
            filename=f"{filename_prefix}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=buffer.getvalue(),
        )
